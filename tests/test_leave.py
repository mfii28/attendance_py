"""Unit tests for leave tracking."""

import sys
import tempfile
import unittest
from datetime import date, time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import Config
from database import Database
from leave_importer import LeaveImporter
from leave_manager import LeaveManager, annual_balance, monthly_absences
from report_generator import ReportGenerator


class TestLeaveLogic(unittest.TestCase):
    def test_deduction_weights(self):
        self.assertEqual(monthly_absences(["H"]), 1.0)
        self.assertEqual(monthly_absences(["H1"]), 0.5)
        self.assertEqual(monthly_absences(["Q"]), 0.25)
        self.assertEqual(monthly_absences(["H1", "H2", "Q"]), 1.25)

    def test_annual_balance(self):
        result = annual_balance(33, 3, [2.0, 1.5, 0.5])
        self.assertEqual(result["entitlement"], 36)
        self.assertEqual(result["used"], 4.0)
        self.assertEqual(result["balance"], 32.0)


class TestLeaveDatabase(unittest.TestCase):
    def setUp(self):
        self.tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp_db.close()
        self.db = Database(Path(self.tmp_db.name))
        self.config = Config()
        self.leave_mgr = LeaveManager(self.db, self.config)
        self.emp_id = self.db.get_or_create_employee("20", "Patience Tsikudo")

    def tearDown(self):
        Path(self.tmp_db.name).unlink(missing_ok=True)

    def test_leave_crud(self):
        self.db.upsert_leave_record(self.emp_id, date(2026, 1, 5), "V")
        records = self.db.get_leave_for_employee(
            self.emp_id, date(2026, 1, 1), date(2026, 1, 31)
        )
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["code"], "V")
        self.db.upsert_leave_record(self.emp_id, date(2026, 1, 5), "S")
        records = self.db.get_leave_for_employee(
            self.emp_id, date(2026, 1, 1), date(2026, 1, 31)
        )
        self.assertEqual(records[0]["code"], "S")
        self.db.delete_leave_record(self.emp_id, date(2026, 1, 5))
        records = self.db.get_leave_for_employee(
            self.emp_id, date(2026, 1, 1), date(2026, 1, 31)
        )
        self.assertEqual(len(records), 0)

    def test_monthly_and_annual_totals(self):
        self.db.set_entitlement(self.emp_id, 2026, 33, 3)
        self.db.upsert_leave_record(self.emp_id, date(2026, 1, 1), "V")
        self.db.upsert_leave_record(self.emp_id, date(2026, 1, 2), "V")
        self.db.upsert_leave_record(self.emp_id, date(2026, 2, 1), "H1")
        totals = self.leave_mgr.compute_annual_leave_totals(self.emp_id, 2026)
        self.assertEqual(totals["entitlement"], 36)
        self.assertEqual(totals["used"], 2.5)
        self.assertEqual(totals["balance"], 33.5)
        self.assertEqual(totals["code_counts"]["V"], 2)
        self.assertEqual(totals["code_counts"]["H1"], 1)

    def test_match_employee_by_name(self):
        matched = self.leave_mgr.match_employee_by_name("Patience Tsikudo")
        self.assertEqual(matched, self.emp_id)
        self.assertIsNone(self.leave_mgr.match_employee_by_name("Unknown Person"))

    def test_attendance_excused_absence(self):
        self.db.upsert_leave_record(self.emp_id, date(2026, 5, 4), "V")
        summaries = self.db.compute_employee_summary(
            date(2026, 5, 1),
            date(2026, 5, 31),
            [0, 1, 2, 3, 4],
            excused_codes=["V", "S", "M", "C", "H", "H1", "H2", "Q", "I"],
        )
        emp_summary = next(s for s in summaries if s["employee_id"] == self.emp_id)
        self.assertEqual(emp_summary.get("excused_days", 0), 1)

    def test_conflict_detection(self):
        self.db.upsert_leave_record(self.emp_id, date(2026, 5, 3), "V")
        self.db.upsert_attendance(
            self.emp_id, date(2026, 5, 3), time(8, 0), time(17, 0), False, 0, "test.csv"
        )
        conflicts = self.leave_mgr.get_conflicts(self.emp_id, date(2026, 5, 3))
        self.assertTrue(any("punch exists" in c for c in conflicts))


class TestLeaveImportExport(unittest.TestCase):
    def setUp(self):
        self.tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp_db.close()
        self.db = Database(Path(self.tmp_db.name))
        self.config = Config()
        self.config._data["paths"]["exports"] = str(Path(tempfile.mkdtemp()))
        self.leave_mgr = LeaveManager(self.db, self.config)
        self.importer = LeaveImporter(self.db, self.config)
        self.report_gen = ReportGenerator(self.db, self.config)
        self.workbook = Path(__file__).resolve().parent.parent / "Leave Tracker 2026.xlsx"

    def tearDown(self):
        Path(self.tmp_db.name).unlink(missing_ok=True)

    def test_export_structure(self):
        emp_id = self.db.get_or_create_employee("20", "Patience Tsikudo")
        self.db.upsert_leave_record(emp_id, date(2026, 6, 1), "V")
        paths = self.report_gen.generate(
            "leave_workbook", date(2026, 1, 1), date(2026, 12, 31), "excel"
        )
        self.assertEqual(len(paths), 1)
        self.assertTrue(paths[0].exists())
        from openpyxl import load_workbook

        wb = load_workbook(paths[0], read_only=True)
        self.assertEqual(len(wb.sheetnames), 13)
        self.assertIn("Totals", wb.sheetnames)
        self.assertIn("January 2026", wb.sheetnames)
        wb.close()

    @unittest.skipUnless(
        Path(__file__).resolve().parent.parent.joinpath("Leave Tracker 2026.xlsx").exists(),
        "Sample workbook not present",
    )
    def test_import_sample_workbook(self):
        for name in (
            "Patience Tsikudo",
            "Emmanuel Tetteh",
            "Solomon Sefah",
            "Maame Lankar De Veer",
        ):
            self.db.get_or_create_employee(str(hash(name) % 10000), name)
        meta, _ = self.importer.preview_file(self.workbook)
        self.assertTrue(meta.get("valid"))
        result = self.importer.import_file(self.workbook, 2026)
        self.assertTrue(result.success)
        self.assertGreater(result.records_imported, 0)


if __name__ == "__main__":
    unittest.main()
