"""Excel import for Leave Tracker workbooks."""

import re
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from config import Config
from database import Database
from leave_manager import (
    MONTH_NAMES,
    LeaveManager,
    month_date_range,
)
from utils import effective_display_name, logger


class LeaveImportResult:
    """Result of a leave workbook import."""

    def __init__(self):
        self.success = False
        self.records_imported = 0
        self.entitlements_imported = 0
        self.records_skipped = 0
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.matched_employees: List[Dict[str, Any]] = []
        self.unmatched_names: List[str] = []


class LeaveImporter:
    """Import Leave Tracker Excel workbooks."""

    MONTH_SHEET_PATTERN = re.compile(
        r"^(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})$"
    )
    DATA_START_ROW = 6
    NAME_COL = 2
    DAY_START_COL = 3
    DAY_END_COL = 33

    def __init__(self, db: Database, config: Config):
        self.db = db
        self.config = config
        self.leave_mgr = LeaveManager(db, config)

    def preview_file(self, file_path: Path) -> Tuple[Dict[str, Any], List[str]]:
        """Analyze workbook without writing to database."""
        meta: Dict[str, Any] = {
            "valid": False,
            "year": None,
            "month_sheets": [],
            "has_totals": False,
            "employee_count": 0,
            "record_estimate": 0,
        }
        warnings: List[str] = []
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            meta["validation_message"] = str(exc)
            return meta, warnings

        years = set()
        for sheet_name in wb.sheetnames:
            if sheet_name == "Totals":
                meta["has_totals"] = True
                continue
            match = self.MONTH_SHEET_PATTERN.match(sheet_name.strip())
            if match:
                month_name, year_str = match.groups()
                years.add(int(year_str))
                meta["month_sheets"].append(sheet_name)

        if not meta["month_sheets"]:
            meta["validation_message"] = "No monthly sheets found (expected 'January YYYY' format)."
            wb.close()
            return meta, warnings

        if len(years) == 1:
            meta["year"] = years.pop()
        else:
            meta["year"] = max(years) if years else None
            warnings.append(f"Multiple years detected: {sorted(years)}")

        unmatched: set = set()
        matched: set = set()
        record_estimate = 0
        for sheet_name in meta["month_sheets"]:
            ws = wb[sheet_name]
            for row in ws.iter_rows(
                min_row=self.DATA_START_ROW, max_col=self.DAY_END_COL, values_only=True
            ):
                if not row or len(row) < self.NAME_COL:
                    continue
                name = row[self.NAME_COL - 1]
                if not name or not str(name).strip():
                    continue
                name_str = str(name).strip()
                emp_id = self.leave_mgr.match_employee_by_name(name_str)
                if emp_id:
                    matched.add(emp_id)
                else:
                    unmatched.add(name_str)
                for col_idx in range(self.DAY_START_COL - 1, min(len(row), self.DAY_END_COL)):
                    val = row[col_idx]
                    if val and self.leave_mgr.is_valid_leave_code(str(val).strip()):
                        record_estimate += 1

        meta["employee_count"] = len(matched)
        meta["record_estimate"] = record_estimate
        meta["matched_count"] = len(matched)
        meta["unmatched_count"] = len(unmatched)
        meta["unmatched_names"] = sorted(unmatched)[:20]
        meta["valid"] = True
        wb.close()
        if unmatched:
            warnings.append(
                f"{len(unmatched)} employee name(s) could not be matched to the database."
            )
        return meta, warnings

    def import_file(self, file_path: Path, year: Optional[int] = None) -> LeaveImportResult:
        result = LeaveImportResult()
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as exc:
            result.errors.append(str(exc))
            return result

        import_year = year
        for sheet_name in wb.sheetnames:
            match = self.MONTH_SHEET_PATTERN.match(sheet_name.strip())
            if match:
                if import_year is None:
                    import_year = int(match.group(2))
                self._import_month_sheet(wb[sheet_name], import_year, result)

        if "Totals" in wb.sheetnames and import_year:
            self._import_totals_sheet(wb["Totals"], import_year, result)

        wb.close()
        result.success = not result.errors
        logger.info(
            "Leave import: %d records, %d entitlements, %d warnings",
            result.records_imported,
            result.entitlements_imported,
            len(result.warnings),
        )
        return result

    def _import_month_sheet(
        self, ws, year: int, result: LeaveImportResult
    ) -> None:
        match = self.MONTH_SHEET_PATTERN.match(ws.title.strip())
        if not match:
            return
        month_name = match.group(1)
        month = MONTH_NAMES.index(month_name) + 1
        _, end = month_date_range(year, month)
        days_in_month = end.day

        for row_idx in range(self.DATA_START_ROW, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=self.NAME_COL).value
            if not name_cell or not str(name_cell).strip():
                continue
            name_str = str(name_cell).strip()
            emp_id = self.leave_mgr.match_employee_by_name(name_str)
            if not emp_id:
                if name_str not in result.unmatched_names:
                    result.unmatched_names.append(name_str)
                    result.warnings.append(f"Unmatched employee: {name_str}")
                continue

            emp = self.db.get_employee(emp_id)
            if emp and emp_id not in {m["employee_id"] for m in result.matched_employees}:
                result.matched_employees.append({
                    "employee_id": emp_id,
                    "name": name_str,
                    "display_name": effective_display_name(emp),
                    "enno": emp["enno"],
                })

            for day in range(1, days_in_month + 1):
                col = self.DAY_START_COL + day - 1
                val = ws.cell(row=row_idx, column=col).value
                if val is None:
                    continue
                code = str(val).strip()
                if not code or not self.leave_mgr.is_valid_leave_code(code):
                    continue
                record_date = date(year, month, day)
                self.db.upsert_leave_record(emp_id, record_date, code)
                result.records_imported += 1

    def _import_totals_sheet(self, ws, year: int, result: LeaveImportResult) -> None:
        for row_idx in range(4, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=2).value
            if not name_cell or not str(name_cell).strip():
                continue
            name_str = str(name_cell).strip()
            emp_id = self.leave_mgr.match_employee_by_name(name_str)
            if not emp_id:
                continue
            alloc = ws.cell(row=row_idx, column=3).value
            carry = ws.cell(row=row_idx, column=4).value
            try:
                annual = float(alloc) if alloc is not None else 0.0
                carry_over = float(carry) if carry is not None else 0.0
            except (TypeError, ValueError):
                continue
            self.db.set_entitlement(emp_id, year, annual, carry_over)
            result.entitlements_imported += 1
