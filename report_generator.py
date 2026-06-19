"""PDF and Excel report generation."""

import html
import json
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from config import Config
from database import Database
from leave_manager import (
    LEAVE_TYPE_ORDER,
    MONTH_ABBREV,
    MONTH_NAMES,
    LeaveManager,
    month_date_range,
)
from utils import effective_display_name, format_date, logger, normalize_enno


class ReportGenerator:
    """Generate PDF and Excel attendance reports."""

    def __init__(self, db: Database, config: Config):
        self.db = db
        self.config = config
        self._output_dir = config.get_export_dir()

    def _export_path(self, filename: str) -> Path:
        return self._output_dir / filename

    def generate(
        self,
        report_type: str,
        start: date,
        end: date,
        fmt: str = "pdf",
        include_charts: bool = True,
        include_raw: bool = False,
        output_dir: Optional[Path] = None,
    ) -> List[Path]:
        """Generate report(s) and return file paths."""
        export_dir = Path(output_dir) if output_dir else self.config.get_export_dir()
        export_dir.mkdir(parents=True, exist_ok=True)
        self._output_dir = export_dir
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        paths = []

        if report_type in ("leave_workbook", "leave_month"):
            if fmt in ("excel", "both"):
                if report_type == "leave_workbook":
                    paths.append(self._leave_workbook_excel(start.year, stamp))
                else:
                    paths.append(
                        self._leave_month_excel(start.year, start.month, stamp)
                    )
            logger.info("Generated %d leave report(s): %s", len(paths), report_type)
            return paths

        if fmt in ("pdf", "both"):
            if report_type == "executive":
                paths.append(self._executive_pdf(start, end, stamp, include_raw))
            elif report_type == "late_offenders":
                paths.append(self._late_offenders_pdf(start, end, stamp))
            elif report_type == "perfect_attendance":
                paths.append(self._perfect_attendance_pdf(start, end, stamp))
            else:
                paths.append(self._executive_pdf(start, end, stamp, include_raw))

        if fmt in ("excel", "both"):
            paths.append(self._excel_report(start, end, stamp, include_raw))

        if fmt in ("html", "both"):
            paths.append(self._html_dashboard(start, end, stamp))

        logger.info("Generated %d report(s): %s", len(paths), report_type)
        return paths

    def _get_summaries(self, start: date, end: date) -> List[Dict]:
        working_days = self.config.get("attendance", "working_days", [0, 1, 2, 3, 4])
        holidays = [
            datetime.strptime(h, "%Y-%m-%d").date()
            for h in self.config.get("attendance", "holidays", [])
        ]
        excused = self.config.get("leave", "excused_codes", [])
        return self.db.compute_employee_summary(
            start, end, working_days, holidays, excused_codes=excused
        )

    def _get_kpi_params(self, start: date, end: date) -> tuple:
        working_days = self.config.get("attendance", "working_days", [0, 1, 2, 3, 4])
        holidays = [
            datetime.strptime(h, "%Y-%m-%d").date()
            for h in self.config.get("attendance", "holidays", [])
        ]
        excused = self.config.get("leave", "excused_codes", [])
        return working_days, holidays, excused

    def _executive_pdf(
        self, start: date, end: date, stamp: str, include_raw: bool
    ) -> Path:
        path = self._export_path(f"Executive_Summary_{stamp}.pdf")
        summaries = self._get_summaries(start, end)
        working_days, holidays, excused = self._get_kpi_params(start, end)
        kpis = self.db.get_dashboard_kpis(start, end, working_days, holidays, excused)
        offenders = self.db.get_top_offenders(start, end, 5, working_days, holidays, excused)

        doc = SimpleDocTemplate(str(path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        logo_path = self.config.assets_dir / "logo.png"
        if self.config.get("export", "include_logo_in_pdf", True) and logo_path.exists():
            story.append(Image(str(logo_path), width=2 * inch, height=1 * inch))
            story.append(Spacer(1, 12))

        title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=18, spaceAfter=20)
        story.append(Paragraph("Executive Attendance Summary", title_style))
        story.append(Paragraph(
            f"Period: {format_date(start)} to {format_date(end)}",
            styles["Normal"],
        ))
        story.append(Spacer(1, 20))

        kpi_data = [
            ["Metric", "Value"],
            ["Active Employees", str(kpis["active_employees"])],
            ["Overall Attendance Rate", f"{kpis['attendance_rate']:.1f}%"],
            ["Total Late Incidents", str(kpis["late_incidents"])],
            ["Perfect Attendance", str(kpis["perfect_attendance"])],
        ]
        kpi_table = Table(kpi_data, colWidths=[3 * inch, 2 * inch])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2196F3")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 20))

        if summaries:
            emp_data = [["#", "ID", "Name", "Present", "Absent", "Att %", "Late", "Late %"]]
            for i, s in enumerate(summaries, 1):
                emp_data.append([
                    str(i), normalize_enno(s["enno"]), s["display_name"][:25],
                    str(s["days_present"]), str(s["absent_days"]),
                    f"{s['attendance_rate']:.1f}%",
                    str(s["late_days"]), f"{s['lateness_rate']:.1f}%",
                ])
            emp_table = Table(emp_data, repeatRows=1)
            emp_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(Paragraph("Employee Summary", styles["Heading2"]))
            story.append(emp_table)

        if offenders:
            story.append(Spacer(1, 20))
            story.append(Paragraph("Top Late Offenders", styles["Heading2"]))
            off_data = [["Name", "Late Days", "Lateness %"]]
            for o in offenders:
                off_data.append([o["display_name"], str(o["late_days"]), f"{o['lateness_rate']:.1f}%"])
            off_table = Table(off_data)
            off_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF5722")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(off_table)

        if include_raw and summaries:
            story.append(PageBreak())
            story.append(Paragraph("Raw Attendance Data", styles["Heading2"]))
            raw = self.db.get_all_attendance(start, end)
            if raw:
                raw_data = [["Date", "ID", "Name", "Arrival", "Departure", "Late"]]
                for r in raw[:500]:
                    raw_data.append([
                        r["date"], normalize_enno(r["enno"]), effective_display_name(r)[:20],
                        r.get("arrival_time") or "-", r.get("departure_time") or "-",
                        "Yes" if r["is_late"] else "No",
                    ])
                raw_table = Table(raw_data, repeatRows=1)
                raw_table.setStyle(TableStyle([
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                story.append(raw_table)
            else:
                story.append(Paragraph("No data available for this period.", styles["Normal"]))

        doc.build(story)
        return path

    def _late_offenders_pdf(self, start: date, end: date, stamp: str) -> Path:
        path = self._export_path(f"Late_Offenders_{stamp}.pdf")
        summaries = self._get_summaries(start, end)
        offenders = sorted(
            [s for s in summaries if s["late_days"] > 0],
            key=lambda x: x["late_days"],
            reverse=True,
        )

        doc = SimpleDocTemplate(str(path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = [
            Paragraph("Late Offenders Report", styles["Heading1"]),
            Paragraph(f"Period: {format_date(start)} to {format_date(end)}", styles["Normal"]),
            Spacer(1, 20),
        ]

        if offenders:
            data = [["#", "Name", "Late Days", "Lateness %", "Avg Arrival"]]
            for i, o in enumerate(offenders, 1):
                data.append([
                    str(i), o["display_name"], str(o["late_days"]),
                    f"{o['lateness_rate']:.1f}%", o["avg_arrival"],
                ])
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F44336")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No late incidents in this period.", styles["Normal"]))

        doc.build(story)
        return path

    def _perfect_attendance_pdf(self, start: date, end: date, stamp: str) -> Path:
        path = self._export_path(f"Perfect_Attendance_{stamp}.pdf")
        summaries = self._get_summaries(start, end)
        perfect = [s for s in summaries if s["late_days"] == 0 and s["days_present"] > 0]

        doc = SimpleDocTemplate(str(path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = [
            Paragraph("Perfect Attendance Report", styles["Heading1"]),
            Paragraph(f"Period: {format_date(start)} to {format_date(end)}", styles["Normal"]),
            Spacer(1, 20),
        ]

        if perfect:
            data = [["#", "Name", "Days Present", "Attendance %"]]
            for i, p in enumerate(perfect, 1):
                data.append([str(i), p["display_name"], str(p["days_present"]), f"{p['attendance_rate']:.1f}%"])
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No perfect attendance records in this period.", styles["Normal"]))

        doc.build(story)
        return path

    def _excel_report(self, start: date, end: date, stamp: str, include_raw: bool) -> Path:
        path = self._export_path(f"Attendance_Report_{stamp}.xlsx")
        summaries = self._get_summaries(start, end)
        working_days, holidays, excused = self._get_kpi_params(start, end)
        kpis = self.db.get_dashboard_kpis(start, end, working_days, holidays, excused)
        wb = Workbook()

        # Sheet 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1["A1"] = "Attendance Executive Summary"
        ws1["A1"].font = Font(bold=True, size=14)
        ws1["A2"] = f"Period: {format_date(start)} to {format_date(end)}"
        summary_rows = [
            ("Active Employees", kpis["active_employees"]),
            ("Overall Attendance Rate", f"{kpis['attendance_rate']}%"),
            ("Total Late Incidents", kpis["late_incidents"]),
            ("Perfect Attendance", kpis["perfect_attendance"]),
        ]
        for i, (label, val) in enumerate(summary_rows, 4):
            ws1[f"A{i}"] = label
            ws1[f"B{i}"] = val

        # Sheet 2: Employee Summary
        ws2 = wb.create_sheet("Employee Summary")
        headers = ["#", "Employee ID", "Name", "Days Present", "Absent Days", "Attendance %", "Late Days", "Lateness %"]
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        for i, s in enumerate(summaries, 1):
            ws2.append([
                i, normalize_enno(s["enno"]), s["display_name"], s["days_present"], s["absent_days"],
                s["attendance_rate"], s["late_days"], s["lateness_rate"],
            ])

        # Sheet 3: Raw Data
        ws3 = wb.create_sheet("Raw Attendance Data")
        raw_headers = ["Date", "Employee ID", "Name", "Arrival", "Departure", "Late", "Late Minutes"]
        ws3.append(raw_headers)
        if include_raw:
            raw = self.db.get_all_attendance(start, end)
            for r in raw:
                ws3.append([
                    r["date"], normalize_enno(r["enno"]), effective_display_name(r),
                    r.get("arrival_time"), r.get("departure_time"),
                    "Yes" if r["is_late"] else "No", r.get("late_minutes", 0),
                ])

        # Sheet 4: Pivot-ready aggregated data
        ws4 = wb.create_sheet("Aggregated Data")
        ws4.append(["Employee ID", "Name", "Month", "Days Present", "Late Days", "Attendance Rate"])
        metrics = self.db.get_monthly_metrics(12)
        for m in metrics:
            month_start = date(m["year"], m["month"], 1)
            if m["month"] == 12:
                month_end = date(m["year"], 12, 31)
            else:
                month_end = date(m["year"], m["month"] + 1, 1) - timedelta(days=1)
            month_label = f"{m['year']}-{m['month']:02d}"
            for s in self._get_summaries(max(start, month_start), min(end, month_end)):
                ws4.append([
                    normalize_enno(s["enno"]), s["display_name"], month_label,
                    s["days_present"], s["late_days"], s["attendance_rate"],
                ])

        wb.save(path)
        return path

    def _badge_class(self, attendance_rate: int) -> str:
        if attendance_rate >= 90:
            return "badge-success"
        if attendance_rate >= 75:
            return "badge-warning"
        return "badge-danger"

    def _html_dashboard(self, start: date, end: date, stamp: str) -> Path:
        """Generate interactive HTML dashboard matching the bash visualizer output."""
        path = self._export_path(f"attendance_dashboard_{stamp}.html")
        summaries = self._get_summaries(start, end)
        working_days, holidays, excused = self._get_kpi_params(start, end)
        kpis = self.db.get_dashboard_kpis(start, end, working_days, holidays, excused)

        names = [s["display_name"] for s in summaries]
        late_days = [s["late_days"] for s in summaries]
        attendance_rates = [s["attendance_rate"] for s in summaries]
        lateness_rates = [s["lateness_rate"] for s in summaries]

        table_rows = []
        for i, s in enumerate(summaries, 1):
            badge = self._badge_class(s["attendance_rate"])
            table_rows.append(
                f"<tr><td>{i}</td><td>{html.escape(normalize_enno(s['enno']))}</td>"
                f"<td>{html.escape(s['display_name'])}</td><td>{s['days_present']}</td>"
                f"<td>{s['absent_days']}</td>"
                f"<td><span class=\"badge {badge}\">{s['attendance_rate']}%</span></td>"
                f"<td>{s['late_days']}</td><td>{s['lateness_rate']}%</td></tr>"
            )

        page = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Attendance Dashboard - Weekdays Only</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        .container {{ max-width: 1400px; margin: 0 auto; }}
        h1 {{
            color: white; text-align: center; margin-bottom: 10px;
            font-size: 2.5em; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        }}
        .subtitle {{
            color: rgba(255,255,255,0.9); text-align: center; margin-bottom: 30px;
        }}
        .stats-grid {{
            display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px; margin-bottom: 30px;
        }}
        .stat-card {{
            background: white; border-radius: 15px; padding: 20px; text-align: center;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        }}
        .stat-number {{ font-size: 2.5em; font-weight: bold; color: #2a5298; }}
        .stat-label {{ color: #666; margin-top: 5px; font-size: 0.9em; }}
        .chart-grid {{
            display: grid; grid-template-columns: repeat(auto-fit, minmax(500px, 1fr));
            gap: 25px; margin-bottom: 25px;
        }}
        .chart-card {{
            background: white; border-radius: 15px; padding: 20px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        }}
        .chart-card h3 {{
            color: #2a5298; margin-bottom: 15px;
            border-left: 4px solid #2a5298; padding-left: 12px;
        }}
        canvas {{ max-height: 400px; width: 100%; }}
        .table-container {{
            background: white; border-radius: 15px; padding: 20px; margin-top: 25px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2); overflow-x: auto;
        }}
        table {{ width: 100%; border-collapse: collapse; }}
        th {{ background: #2a5298; color: white; padding: 12px; text-align: left; }}
        td {{ padding: 10px 12px; border-bottom: 1px solid #e0e0e0; }}
        tr:hover {{ background: #f5f5f5; }}
        .badge {{
            display: inline-block; padding: 4px 10px; border-radius: 20px;
            font-size: 0.85em; font-weight: bold;
        }}
        .badge-success {{ background: #d4edda; color: #155724; }}
        .badge-warning {{ background: #fff3cd; color: #856404; }}
        .badge-danger {{ background: #f8d7da; color: #721c24; }}
        .footer {{
            text-align: center; color: rgba(255,255,255,0.7); margin-top: 30px; padding: 20px;
        }}
        .note {{
            background: #fff3cd; border-left: 4px solid #ffc107; padding: 10px;
            margin-bottom: 20px; border-radius: 5px; font-size: 0.9em;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Attendance Analytics Dashboard</h1>
        <div class="subtitle">Late Threshold: 08:01 AM | Weekends Excluded | Period: {format_date(start)} to {format_date(end)}</div>
        <div class="note">
            <strong>Note:</strong> Saturdays and Sundays are EXCLUDED from working days. Only Monday-Friday considered.
            Working days = unique weekday dates with scan records ({kpis.get('working_days', 0)} days).
        </div>
        <div class="stats-grid">
            <div class="stat-card"><div class="stat-number">{kpis['active_employees']}</div><div class="stat-label">Total Employees</div></div>
            <div class="stat-card"><div class="stat-number">{kpis.get('working_days', 0)}</div><div class="stat-label">Working Days (Mon-Fri)</div></div>
            <div class="stat-card"><div class="stat-number">{kpis['late_incidents']}</div><div class="stat-label">Total Late Incidents</div></div>
            <div class="stat-card"><div class="stat-number">{kpis['perfect_attendance']}</div><div class="stat-label">Perfect Punctuality</div></div>
        </div>
        <div class="chart-grid">
            <div class="chart-card"><h3>Late Days by Employee</h3><canvas id="lateChart"></canvas></div>
            <div class="chart-card"><h3>Lateness Rate (%)</h3><canvas id="latenessRateChart"></canvas></div>
            <div class="chart-card"><h3>Attendance Rate (%)</h3><canvas id="attendanceChart"></canvas></div>
            <div class="chart-card"><h3>Performance Distribution</h3><canvas id="distributionChart"></canvas></div>
        </div>
        <div class="table-container">
            <h3 style="margin-bottom: 15px;">Employee Attendance Register (Weekdays Only)</h3>
            <table>
                <thead>
                    <tr><th>#</th><th>Employee ID</th><th>Name</th><th>Days Present</th>
                    <th>Absent Days</th><th>Attendance %</th><th>Late Days</th><th>Lateness %</th></tr>
                </thead>
                <tbody>{''.join(table_rows)}</tbody>
            </table>
        </div>
        <div class="footer">
            Generated by Attendance Management System | Late = arrival at or after 08:01 AM | Weekends excluded
        </div>
    </div>
    <script>
        const employees = {json.dumps(names)};
        const lateDays = {json.dumps(late_days)};
        const attendanceRates = {json.dumps(attendance_rates)};
        const latenessRates = {json.dumps(lateness_rates)};

        new Chart(document.getElementById('lateChart'), {{
            type: 'bar',
            data: {{
                labels: employees,
                datasets: [{{
                    label: 'Late Days',
                    data: lateDays,
                    backgroundColor: 'rgba(220, 53, 69, 0.7)',
                    borderColor: 'rgba(220, 53, 69, 1)',
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                scales: {{ y: {{ beginAtZero: true, title: {{ display: true, text: 'Number of Late Days' }} }} }}
            }}
        }});

        new Chart(document.getElementById('latenessRateChart'), {{
            type: 'bar',
            data: {{
                labels: employees,
                datasets: [{{
                    label: 'Lateness Rate (%)',
                    data: latenessRates,
                    backgroundColor: 'rgba(255, 193, 7, 0.7)',
                    borderColor: 'rgba(255, 193, 7, 1)',
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                scales: {{ y: {{ max: 100, title: {{ display: true, text: 'Percentage (%)' }} }} }}
            }}
        }});

        new Chart(document.getElementById('attendanceChart'), {{
            type: 'bar',
            data: {{
                labels: employees,
                datasets: [{{
                    label: 'Attendance Rate (%)',
                    data: attendanceRates,
                    backgroundColor: 'rgba(40, 167, 69, 0.7)',
                    borderColor: 'rgba(40, 167, 69, 1)',
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                scales: {{ y: {{ max: 100, title: {{ display: true, text: 'Percentage (%)' }} }} }}
            }}
        }});

        const zeroLate = latenessRates.filter(r => r === 0).length;
        const lowLate = latenessRates.filter(r => r > 0 && r <= 20).length;
        const medLate = latenessRates.filter(r => r > 20 && r <= 40).length;
        const highLate = latenessRates.filter(r => r > 40).length;

        new Chart(document.getElementById('distributionChart'), {{
            type: 'pie',
            data: {{
                labels: ['0% Lateness', '1-20% Lateness', '21-40% Lateness', '40%+ Lateness'],
                datasets: [{{
                    data: [zeroLate, lowLate, medLate, highLate],
                    backgroundColor: ['#28a745', '#ffc107', '#fd7e14', '#dc3545']
                }}]
            }},
            options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom' }} }} }}
        }});
    </script>
</body>
</html>"""
        path.write_text(page, encoding="utf-8")
        logger.info("HTML dashboard written to %s", path)
        return path

    def _leave_deduction_formula(self, row: int) -> str:
        day_range = f"C{row}:AG{row}"
        parts = [
            f'(COUNTIF({day_range},"H")*1)',
            f'(COUNTIF({day_range},"H1")*0.5)',
            f'(COUNTIF({day_range},"H2")*0.5)',
            f'(COUNTIF({day_range},"Q")*0.25)',
            f'(COUNTIF({day_range},"I")*1)',
            f'(COUNTIF({day_range},"S")*1)',
            f'(COUNTIF({day_range},"M")*1)',
            f'(COUNTIF({day_range},"C")*1)',
            f'(COUNTIF({day_range},"V")*1)',
            f'(COUNTIF({day_range},"A")*1)',
        ]
        return "=" + "+".join(parts)

    def _write_leave_month_sheet(
        self, ws, year: int, month: int, employees: List[Dict], data_start: int = 6
    ) -> None:
        _, end = month_date_range(year, month)
        days_in_month = end.day
        month_label = MONTH_ABBREV[month - 1]

        ws["C1"] = year
        ws["C3"] = month_label
        for col in range(3, 3 + days_in_month):
            ws.cell(row=4, column=col, value=col - 2)
        ws["AH3"] = "Absences this month"
        ws["AJ3"] = "Absence type"
        ws["AK3"] = "Code"
        ws["AL3"] = "Deduction"

        for col in range(3, 3 + days_in_month):
            day = col - 2
            d = date(year, month, day)
            ws.cell(row=5, column=col, value=d.strftime("%a")[:3])

        leave_mgr = LeaveManager(self.db, self.config)
        grid = leave_mgr.get_month_grid(year, month)
        by_id = {e["employee_id"]: e for e in grid["employees"]}

        for i, emp in enumerate(employees):
            row = data_start + i
            ws.cell(row=row, column=1, value=i + 1)
            ws.cell(row=row, column=2, value=emp["export_name"])
            emp_grid = by_id.get(emp["id"], {})
            days = emp_grid.get("days", {})
            for day in range(1, days_in_month + 1):
                code = days.get(day)
                if code:
                    ws.cell(row=row, column=2 + day, value=code)
            ws.cell(row=row, column=34, value=self._leave_deduction_formula(row))

    def _leave_workbook_excel(self, year: int, stamp: str) -> Path:
        path = self._export_path(f"Leave_Tracker_{year}_{stamp}.xlsx")
        employees = self.db.get_all_employees(active_only=True)
        wb = Workbook()
        wb.remove(wb.active)

        for month in range(1, 13):
            title = f"{MONTH_NAMES[month - 1]} {year}"
            ws = wb.create_sheet(title)
            self._write_leave_month_sheet(ws, year, month, employees)

        ws_totals = wb.create_sheet("Totals")
        ws_totals["C3"] = "Allocation"
        ws_totals["D3"] = "Carry over"
        ws_totals["E3"] = "Total"
        ws_totals["G3"] = "Used"
        ws_totals["H3"] = "Balance"
        type_headers = {
            "H": "Holiday",
            "H1": "Half Day (morning)",
            "H2": "Half Day (afternoon)",
            "Q": "Quarter day",
            "I": "Inactive",
            "S": "Sickness",
            "M": "Maternity/Paternity",
            "C": "Compassionate",
            "V": "Vacation",
            "A": "Absent/No Show",
        }
        for j, code in enumerate(LEAVE_TYPE_ORDER):
            ws_totals.cell(row=3, column=10 + j, value=type_headers.get(code, code))

        leave_mgr = LeaveManager(self.db, self.config)
        for i, emp in enumerate(employees):
            row = 4 + i
            ws_totals.cell(row=row, column=1, value=i + 1)
            ws_totals.cell(row=row, column=2, value=emp["export_name"])
            totals = leave_mgr.compute_annual_leave_totals(emp["id"], year)
            ws_totals.cell(row=row, column=3, value=totals["annual_allocation"])
            ws_totals.cell(row=row, column=4, value=totals["carry_over"])
            ws_totals.cell(row=row, column=5, value=f"=SUM(C{row}:D{row})")
            ah_refs = [
                f"'{MONTH_NAMES[m - 1]} {year}'!AH{6 + i}" for m in range(1, 13)
            ]
            ws_totals.cell(row=row, column=7, value=f"=SUM({','.join(ah_refs)})")
            ws_totals.cell(row=row, column=8, value=f"=E{row}-G{row}")
            for j, code in enumerate(LEAVE_TYPE_ORDER):
                count_refs = []
                for m in range(1, 13):
                    sheet = f"'{MONTH_NAMES[m - 1]} {year}'"
                    emp_row = 6 + i
                    count_refs.append(
                        f"COUNTIF({sheet}!C{emp_row}:AG{emp_row},\"{code}\")"
                    )
                ws_totals.cell(
                    row=row, column=10 + j, value=f"=SUM({','.join(count_refs)})"
                )

        wb.save(path)
        return path

    def _leave_month_excel(self, year: int, month: int, stamp: str) -> Path:
        path = self._export_path(
            f"Leave_{MONTH_NAMES[month - 1]}_{year}_{stamp}.xlsx"
        )
        employees = self.db.get_all_employees(active_only=True)
        wb = Workbook()
        ws = wb.active
        ws.title = f"{MONTH_NAMES[month - 1]} {year}"
        self._write_leave_month_sheet(ws, year, month, employees)
        wb.save(path)
        return path
